from openpyxl import load_workbook

# Carregar a planilha .xlsx
wb = load_workbook('sua_planilha.xlsx')
sheet = wb.active

# Criar uma lista para armazenar os dados
dados = []
for row in sheet.iter_rows(values_only=True):
    dados.append(row)

# Remover duplicatas (com base em todas as colunas)
dados_unicos = []
for row in dados:
    if row not in dados_unicos:
        dados_unicos.append(row)

# Criar uma nova planilha e salvar
wb_novo = load_workbook('sua_planilha.xlsx')
sheet_novo = wb_novo.active

# Limpar dados antigos e adicionar dados únicos
for row in sheet_novo.iter_rows(min_row=2, max_col=len(dados_unicos[0]), max_row=len(dados_unicos)):
    for cell in row:
        cell.value = None

for row in dados_unicos:
    sheet_novo.append(row)

wb_novo.save('planilha_sem_duplicatas.xlsx')
print("Registros duplicados foram removidos com sucesso.")
